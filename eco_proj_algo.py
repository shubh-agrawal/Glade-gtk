import openpyxl
import math

def magic(numbers):
    return int(''.join([ "%d"%x for x in numbers]))

ite=1
flag=0
totalcost=0

cost=[]
weight=[]
volume=[]
pweight=[]
pvolume=[]
pdemand=[]


out=[[0 for x in range(50)] for y in range(50)]

file_name="inputdata.xlsx"
wb=openpyxl.load_workbook(file_name)
sheet_name_list=wb.get_sheet_names()
wrk_sheet=wb.get_sheet_by_name(sheet_name_list[0])

for x in range(2,8):
	cost.append(wrk_sheet.cell(row=x,column=2).value)

for x in range(2,8):
	weight.append(wrk_sheet.cell(row=x,column=3).value)

for x in range(2,8):
	volume.append(wrk_sheet.cell(row=x,column=4).value)

for x in range(2,41):
	pweight.append(wrk_sheet.cell(row=x,column=7).value)

for x in range(2,41):
	pvolume.append(wrk_sheet.cell(row=x,column=8).value)

for x in range(2,41):
	pdemand.append(wrk_sheet.cell(row=x,column=6).value)

numtruck=[]
for i in range(0,len(pweight)):
	numtruck.append(0.0)

expense=[]
for i in range(0,len(weight)):
	expense.append(0.0)	

dendiff=[]
for i in range(0,len(pweight)):
	dendiff.append(0.0)


den=[]
for i in range(0,len(pweight)):
	den.append(0.0)

q=[]
for i in range(0,len(pweight)):
	q.append(0.0)	

test=[]
for i in range(0,len(pweight)):
	test.append(0.0)		


while flag==0:
	demandweight=0
	demandvolume=0
	slackw=0
	slackv=0

	for i in range(0,len(pweight)):
		den[i]=(pweight[i]/pvolume[i])
		demandweight=demandweight+pdemand[i]*pweight[i]
		demandvolume=demandvolume+pdemand[i]*pvolume[i]

	for i in range(0,len(cost)):
		expense[i]=(cost[i]*(math.trunc(max(demandweight/weight[i],demandvolume/volume[i]))+1))


	costmin=min(expense)
	exp_index=expense.index(costmin)


	qtruck=math.trunc(max(demandweight/weight[i],demandvolume/volume[i]))

	for i in range(0,len(pweight)):
		test[i]=(pdemand[i]/qtruck)
		q[i]=(math.trunc(pdemand[i]/qtruck))
		slackv=slackv+q[i]*pvolume[i]
		slackw=slackw+q[i]*pweight[i]

	#print "test",test
	#print "q",q

	
	slackv=volume[exp_index]-slackv
	slackw=weight[exp_index]-slackw
	slackden=slackw/slackv

	for i in range(0,len(pweight)):
		dendiff[i]=(abs(den[i]-slackden))
		if pdemand[i]==0:
			dendiff[i]=100000
			pweight[i]=10000
			pvolume[i]=100000

	a=min(dendiff);
	slack_index=dendiff.index(a)

	slacknum=math.trunc(min(slackw/pweight[slack_index],slackv/pvolume[slack_index]))
	q[slack_index]=q[slack_index]+slacknum
	slackv=slackv-pvolume[slack_index]*slacknum
	slackw=slackw-pweight[slack_index]*slacknum
	slackden=slackw/slackv
	
	minw=min(pweight)
	minwindex=pweight.index(minw)

	minv=min(pvolume)
	minvindex=pvolume.index(minv)

	if slackv >= minv and slackw>=minw:
		qminv=math.trunc(slackv/minv)
		qminw=math.trunc(slackw/minw)
		slackww=slackw-qminw*pweight[minwindex]
		slackvw=slackv-qminw*pvolume[minwindex]
		slackwv=slackw-qminv*pweight[minvindex]
		slackvv=slackv-qminv*pvolume[minvindex]

		if slackwv>0 and slackvw<0:
			q[minvindex]=q[minvindex]+qminv
			slackw=slackwv
			slackv=slackvv
		
		elif slackwv<0 and slackvw>0:
			q[minwindex]=q[minwindex]+qminw
			slackw=slackww
			slackv=slackvv

		elif slackwv>0 and slackvw>0:
			if abs(slackden-pweight[minwindex]/pvolume[minwindex]) < abs(slackden-pweight[minvindex]/pvolume[minvindex]):
				q[minwindex]=q[minwindex]+qminw
				slackw=slackwv
				slackv=slackvw

			else:
				q[minvindex]=q[minvindex]+qminv
				slackw=slackwv
				slackv=slackvv

		else:
			qminv=min(math.trunc(slackw/pweight[minvindex]),pdemand[minvindex]);
			slackwv=slackw-qminv*pweight[minvindex];
			slackvv=slackv-qminv*pvolume[minvindex];
			qminw=min(math.trunc(slackv/pvolume[minwindex]),pdemand[minwindex]);
			slackww=slackw-qminw*pweight[minwindex];
			slackvw=slackv-qminw*pvolume[minwindex];
			if abs(slackden-pweight[minwindex]/pvolume[minwindex]) < abs(slackden-pweight[minvindex]/pvolume[minvindex]):
				q[minwindex]=q[minwindex]+qminw
				slackw=slackwv
				slackv=slackvw
			else:
				q[minvindex]=q[minvindex]+qminv
				slackw=slackwv
				slackv=slackvv
		

	for i in range(0,len(pweight)):
		if pdemand[i]==0 or q[i]==0:
			numtruck[i]=100000
			continue

		if math.trunc(pdemand[i]/q[i])==0:
			q[i]=pdemand[i]
		numtruck[i]=math.trunc(pdemand[i]/q[i])

	numvalue=min(numtruck)
	numtruckindex=numtruck.index(numvalue)

	for i in range(0,len(pweight)):
		pdemand[i]=pdemand[i]-numvalue*q[i]


	print ite
	out[ite][1]=exp_index

	for i in range(0,len(pweight)):
		out[ite][i+2]=q[i]

	out[ite][2]=numvalue
	
	print "pdemand",pdemand

	for i in range(0,len(pdemand)):
		if pdemand[i]!=0:
			flag=0
			break

		else :
			flag=1 

	#if pdemand==0:
	#	flag=1

	wei=0
	vol=0

	for i in range(0,len(pweight)):
		wei=wei+pweight[i]*out[ite][i+2]
		vol=vol+pvolume[i]*out[ite][i+2]

	out[ite][len(pweight)+3]=1-slackw/weight[exp_index]
	out[ite][len(pweight)+4]=1-slackv/volume[exp_index]

	ite=ite+1
	totalcost=totalcost+cost[exp_index]*numvalue			